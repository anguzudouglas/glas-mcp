import asyncio
import base64
import io
import os
from typing import Any, Dict

from glas_mcp.tools.base import BaseTool

_HEADER_BG = "1A1A2E"
_HEADER_FG = "FFFFFF"


class DocumentCreateXlsxTool(BaseTool):
    """
    MCP tool that converts HTML tables into an XLSX spreadsheet.
    Each <table> becomes a separate worksheet.
    """

    async def execute(self, arguments: Dict[str, Any]) -> Any:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from glas_mcp.tools.document_create_xlsx.helpers.table_parser import parse_tables
        # ── Input validation ────────────────────────────────────────────────
        if not isinstance(arguments, dict):
            return {"error": "Arguments must be a JSON object."}

        html_content: str = arguments.get("html_content", "")
        if not isinstance(html_content, str) or not html_content.strip():
            return {"error": "'html_content' is required and must be a non-empty string."}

        filename: str = str(arguments.get("filename", "spreadsheet.xlsx")).strip()
        output_dir: str = str(arguments.get("output_dir", "./generated_docs/")).strip()
        return_base64: bool = bool(arguments.get("return_base64", True))
        auto_col_width: bool = bool(arguments.get("auto_column_width", True))

        if not filename:
            filename = "spreadsheet.xlsx"
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"

        if not output_dir:
            output_dir = "./generated_docs/"

        # ── Parse HTML tables ───────────────────────────────────────────────
        try:
            tables = parse_tables(html_content)
        except Exception as exc:
            return {"error": f"HTML parsing failed: {exc}"}

        if not tables:
            return {
                "error": (
                    "No <table> elements found in html_content. "
                    "Wrap your data in <table><thead><tr><th>Col</th></tr></thead>"
                    "<tbody><tr><td>Value</td></tr></tbody></table>."
                )
            }

        # ── Create output directory ─────────────────────────────────────────
        try:
            os.makedirs(output_dir, exist_ok=True)
        except OSError as exc:
            return {"error": f"Cannot create output directory '{output_dir}': {exc}"}

        output_path = os.path.join(output_dir, filename)

        # ── Build workbook ──────────────────────────────────────────────────
        try:
            xlsx_bytes = await asyncio.to_thread(
                self._build_xlsx, tables, auto_col_width
            )
        except Exception as exc:
            return {
                "error": f"XLSX generation failed: {exc}",
                "hint": "Check that all table rows have consistent cell counts.",
            }

        # ── Write to disk ───────────────────────────────────────────────────
        try:
            with open(output_path, "wb") as fh:
                fh.write(xlsx_bytes)
        except OSError as exc:
            return {"error": f"Cannot write XLSX to '{output_path}': {exc}"}

        result: Dict[str, Any] = {
            "success": True,
            "file_path": os.path.abspath(output_path),
            "filename": filename,
            "size_bytes": len(xlsx_bytes),
            "sheets": [t["sheet_name"] for t in tables],
            "table_count": len(tables),
        }
        if return_base64:
            result["base64"] = base64.b64encode(xlsx_bytes).decode("utf-8")

        return result

    # ── Private helpers ─────────────────────────────────────────────────────

    @staticmethod
    def _build_xlsx(tables, auto_col_width: bool) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        seen_names: Dict[str, int] = {}

        for table in tables:
            raw_name = table["sheet_name"]
            if raw_name in seen_names:
                seen_names[raw_name] += 1
                sheet_name = f"{raw_name[:28]}_{seen_names[raw_name]}"
            else:
                seen_names[raw_name] = 1
                sheet_name = raw_name

            ws = wb.create_sheet(title=sheet_name)
            col_widths: Dict[int, int] = {}

            # Headers
            headers = table["headers"]
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color=_HEADER_FG)
                    cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
                    cell.alignment = Alignment(horizontal="center")
                    col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(str(header)) + 2)

            # Data rows
            row_offset = 2 if headers else 1
            for row_idx, row in enumerate(table["rows"], start=row_offset):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(str(value)) + 2)

            # Auto column width
            if auto_col_width:
                for col_idx, width in col_widths.items():
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
